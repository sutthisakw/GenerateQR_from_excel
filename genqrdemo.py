# modules needed
import qrcode
from tkinter import filedialog
from tkinter import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# เลือกไฟล์ excel ที่เป็นข้อมูลที่ต้องการ gen QR
# ข้อมูลที่ต้องการ gen จะเริ่มที่คอลัมน์ A แถวที่ 2 เพราะหลังจาก gen จะมีคำสั่งสร้าง header ด้านบน
print('You select file:')
root = Tk()
root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
print (root.filename)

# เลือกโฟลเดอร์ที่เก็บไฟล์ภาพ QR และ excel ใหม่
print('คุณต้องการเซฟไฟล์ excel ที่แปลง QR และไฟล์ภาพ QR ทั้งหมดที่ไหน ?')
root2 = Tk()
root2.withdraw()
folder_selected = filedialog.askdirectory()

# read the excel file
workbook = load_workbook(str(root.filename))
sheet = workbook.active

# settings for qrcode to be produced
qr = qrcode.QRCode(
    version=1,
    error_correction=qrcode.constants.ERROR_CORRECT_L,
    box_size=4,
    border=2,)

# ตั้งค่าขนาดเซลล์ของไฟล์ excel ที่กำลังจะถูกสร้าง 
sheet.column_dimensions['B'].width = 17 #ความกว้างของคอลัมน์
for i in range(1,len(sheet['A'])+1):
    sheet.row_dimensions[i+1].height=110 #ความสูงของแถวแต่ละแถว

# ชื่อหัวคอลัมน์ B
sheet["B1"]="QR_Codes"

# วนลูปสร้าง QR โดยมองจาก A2 เป็นต้นไป ถ้า A ไหนไม่มีข้อมูลจะข้าม
for i in range(2,len(sheet['A'])+1):
    #print(i) #########=> พิมพ์ดูข้อมูลใน terminal
  if sheet.cell(row=i, column=1).value is None: # จะเริ่มเช็คจาก row A column 1 ว่าเป็นค่าว่างหรือไม่ ถ้าว่างให้ข้ามไปดำเนินการการต่อที่ else
    continue 
  else:
    iditem = sheet.cell(row=i, column=1).value
    qr.add_data(iditem)
    #print(iditem)   #########=> พิมพ์ดูข้อมูลใน terminal
    qr.make(fit=True)
    
    img = qrcode.make(sheet.cell(row=i, column=1).value)

    #คำนวณขนาดภาพใหม่
    img_w, img_h = img.size
    ratio = img_w / img_h
    resize_qr = 100 #ระบุขนาดภาพเป็น 100 เพื่อให้การคำนวณเป็น 100x100
    new_h = int(ratio * resize_qr)
    img = img.resize((resize_qr,new_h))
    print('SIZE:',img.resize) #ดูขนาดภาพหลังจากคำนวณใหม่

    # เซฟ qr เป็นภาพ
    img.save(folder_selected + "/" + "no"+str(i)+"_"+iditem +"_qrcode.png")
    
    # สร้าง qr ใน excel
    img=openpyxl.drawing.image.Image(folder_selected + "/" + "no"+str(i)+"_"+iditem +"_qrcode.png")
    print(sheet.cell(row=i,column=1))   ########=> พิมพ์ดูข้อมูลใน terminal
    img.anchor = "B" + str(i)
    sheet.add_image(img)
    sheet["B" + str(i)].alignment = Alignment(horizontal='center', vertical='center')
    sheet["A" + str(i)].alignment = Alignment(horizontal='center', vertical='center')


# สร้างไฟล์ excel เก็บ
workbook.save(folder_selected+ "/qrcode_produced.xlsx")